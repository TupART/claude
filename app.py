from flask import Flask, request, render_template, send_file, jsonify
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename
import os

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def get_pcc_step2_data(windows_username, market, is_pcc, is_ds):
    result = {}
    
    if is_pcc:
        result['name'] = f"{windows_username}.pcc@costa.it"
        result['users'] = f"{windows_username}@es.costa.it"
    
    if is_ds:
        groups_map = {
            'DACH': '180, 130, 97, 160, 29',
            'Italy': '100, 160, 29',
            'Spain': '160, 29',
            'France': '110, 160, 29'
        }
    else:
        groups_map = {
            'DACH': '180, 130, 97',
            'Italy': '100',
            'Spain': '160, 29',
            'France': '110'
        }
    
    result['groups'] = groups_map.get(market, 'No specific groups for this market')
    return result

def get_step3_data(market, is_pcc, is_ds, is_tl):
    department = ''
    colas = ''
    skills = ''
    
    if is_ds:
        if market == 'Spain':
            department = "CCH  Com Ops E B2C"
            colas = "E_CBDsManagement; E_CBDsSales; E_CBCostaClub; E_CBShorexB2C; E_Outbound; E_WAB2C; E_CBWOPT; E_CBWB2C;"
            skills = "DsSales:5, DsManagement:4, CostaClub:5, WA_CostaClub, WA_BookingSales, WA_BookingManagement, Spanish, Barcelona"
        elif market == 'DACH':
            department = "CCH  Com Ops DACH B2C"
            colas = "D_WAB2C, CH_D_WAB2C, A_WAB2C, A_CBDsManagement; A_CBDsSales; A_CBCostaClub; A_Outbound;"
            skills = "DsSales:5, DsManagement:4, CostaClub:5, WA_CostaClub, WA_BookingSales, WA_BookingManagement, German, Barcelona"
        elif market == 'France':
            department = "CCH  Com Ops F B2C"
            colas = "F_CBDsManagement; F_CBDsSales; F_CBCostaClub; F_CBShorexB2C; F_Outbound; F_WAB2C; F_CBWOPT; F_CBWB2B;"
            skills = "DsSales:5, DsManagement:4, CostaClub:5, WA_CostaClub, WA_BookingSales, WA_BookingManagement, French, Barcelona"
    
    elif is_pcc or is_tl:
        if market == 'DACH':
            department = "CCH Com Ops  DACH PCC"
            colas = "A_PCC; A_PCCDefault; A_CBWB2CPCC; A_CBWOPTPCC; A_CBFQPCC; A_CBFQPCCDefault; A_Outbound;"
            skills = "A_PCC:5, D_PCC:5, CH_D_PCC:5, DsSales:1, DsManagement:1, German"
        elif market == 'France':
            department = "CCH  Com Ops  F PCC"
            colas = "F_PCC; F_PCCDefault; F_CBWB2CPCC; F_CBWOPTPCC; F_CBFQPCC; F_CBFQPCCDefault; F_Outbound;"
            skills = "F_PCC, DsSales:1, DsManagement:1, Barcelona, French"
        elif market == 'Spain':
            department = "CCH  Com Ops E PCC"
            colas = "E_PCC; E_PCCDefault; E_Outbound; E_DsManagement; E_CBDsManagement; E_DsSales; E_CBDsSales;"
            skills = "E_PCC:5, DsSales:1, DsManagement:1, Spanish"
        elif market == 'Italy':
            department = "CCH CC ITA PCC"
            colas = "I_PCC; I_PCC_Default; I_CBWB2CPCC; I_CBWOPTPCC; I_CBFQPCC; I_CBFQPCCDefault; I_Outbound;"
            skills = "I_PCC, Barcelona, Italian"
    
    else:
        if market == 'DACH':
            department = "CCH  Com Ops  DACH B2C"
            colas = "A_DsManagement; A_CBDsManagement; A_DsSales; A_CBDsSales; A_CostaClub; A_CBCostaClub;"
            skills = "DsSales:5, DsManagement:4, CostaClub:5, German"
        elif market == 'France':
            department = "CCH  Com Ops  F B2C"
            colas = "F_DsManagement; F_CBDsManagement; F_DsSales; F_CBDsSales; F_CostaClub; F_CBCostaClub;"
            skills = "DsSales:5, DsManagement:4, CostaClub:5, French"
        elif market == 'Spain':
            department = "CCH  Com Ops  E B2C"
            colas = "E_DsManagement; E_CBDsManagement; E_DsSales; E_CBDsSales; E_CostaClub; E_CBCostaClub;"
            skills = "DsSales:5, DsManagement:4, CostaClub:5, Spanish"
    
    return {
        'division': 'UECC',
        'department': department,
        'colas': colas,
        'skills': skills
    }

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            df = pd.read_excel(file_path)
            data = []
            
            for idx, row in df.iterrows():
                name = row['Name']
                surname = row['Surname']
                market = row['Market']
                is_pcc = row['Va a ser PCC?'] == 'Y'
                is_ds = row['Va a ser PCC?'] == 'DS'
                is_tl = row['Va a ser PCC?'] == 'TL'
                windows_username = row['B2E User Name']
                
                step2_data = get_pcc_step2_data(windows_username, market, is_pcc, is_ds)
                step3_data = get_step3_data(market, is_pcc, is_ds, is_tl)
                
                data.append({
                    'id': idx,
                    'name': name,
                    'surname': surname,
                    'step2': step2_data,
                    'step3': step3_data,
                    'step4': {
                        'name': name,
                        'surname': surname,
                        'email': row['E-mail'],
                        'market': market,
                        'pcc_status': row['Va a ser PCC?'],
                        'b2e_username': windows_username
                    }
                })
            
            return render_template('index.html', data=data)
    
    return render_template('index.html')

@app.route('/process_step4', methods=['POST'])
def process_step4():
    selected_rows = request.form.getlist('rows')
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], os.listdir(app.config['UPLOAD_FOLDER'])[0])
    df = pd.read_excel(file_path)
    
    plantilla = 'PlantillaSTEP4.xlsx'
    wb = openpyxl.load_workbook(plantilla)
    ws = wb.active
    
    for row in selected_rows:
        idx = int(row)
        name = df.iloc[idx]['Name']
        surname = df.iloc[idx]['Surname']
        email = df.iloc[idx]['E-mail']
        market = df.iloc[idx]['Market']
        pcc_status = df.iloc[idx]['Va a ser PCC?']
        b2e_username = df.iloc[idx]['B2E User Name']
        
        row_num = 7 + idx
        ws[f'C{row_num}'] = name
        ws[f'D{row_num}'] = surname
        ws[f'E{row_num}'] = email
        
        if pcc_status == 'Y':
            if market == 'DACH':
                ws[f'F{row_num}'] = "/+4940210918145 /+43122709858 /+41445295828"
                ws[f'G{row_num}'] = "D_PCC"
                ws[f'H{row_num}'] = "Team_D_CCH_PCC_1"
            elif market == 'France':
                ws[f'F{row_num}'] = "/+33180037979"
                ws[f'G{row_num}'] = "F_PCC"
                ws[f'H{row_num}'] = "Team_F_CCH_PCC_1"
            elif market == 'Spain':
                ws[f'F{row_num}'] = "/+34932952130"
                ws[f'G{row_num}'] = "E_PCC"
                ws[f'H{row_num}'] = "Team_E_CCH_PCC_1"
            elif market == 'Italy':
                ws[f'F{row_num}'] = "/+390109997099"
                ws[f'G{row_num}'] = "I_PCC"
                ws[f'H{row_num}'] = "Team_I_CCH_PCC_1"
        
        ws[f'L{row_num}'] = "Y" if pcc_status == 'Y' else "N"
        ws[f'Q{row_num}'] = b2e_username
        ws[f'R{row_num}'] = b2e_username
        ws[f'V{row_num}'] = "Team Leader" if pcc_status == 'TL' else "Agent"
    
    output_file = 'PlantillaSTEP4_Rellenada.xlsx'
    wb.save(output_file)
    return send_file(output_file, as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    app.run(debug=True)